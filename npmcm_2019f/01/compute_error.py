import openpyxl
import sys

sys.setrecursionlimit(100000)
alpha1 = 25
alpha2 = 15
beta1 = 20
beta2 = 25
theta = 30
sigma = 0.001
import os

data_path = os.path.join('..', '..', '2019 年中国研究生数学建模竞赛 F 题\附件 1：数据集 1-终
稿.xlsx
')
# 打开excel 文件,获取工作簿对象
wb = openpyxl.load_workbook(data_path)
# 获取指定的表单
sheet = wb['data1']
point_list = []
print(sheet.max_row)
for row in sheet[3:sheet.max_row]:
    point_list.append([row[1].value, row[2].value, row[3].value, row[4].value, row[5].value])
print(point_list)
point_num = len(point_list)


def judge_z(start_index, end_index=point_num - 1):
    x1, y1, z1 = point_list[start_index][0:3]
    x2, y2, z2 = point_list[end_index][0:3]
    if abs(z1 - z2) > 5000:
        return False
    else:
        return True


def get_distance(start_index, end_index):
    x1, y1, z1 = point_list[start_index][0:3]
    x2, y2, z2 = point_list[end_index][0:3]
    return ((x1 - x2) ** 2 + (y1 - y2) ** 2 + (z1 - z2) ** 2) ** 0.5


def judge(start_index, end_index, horizontal_error, vertical_error):
    end_point_type = point_list[end_index][3]
    distance = get_distance(start_index, end_index)


delta_error = distance * sigma
end_point_horizontal_error = horizontal_error + delta_error
end_point_vertical_error = vertical_error + delta_error
if judge_z(start_index):
    if end_index != point_num - 1:  # 下一个点不是终点
        if end_point_type == 0:  # 水平
            if end_point_horizontal_error <= beta2 and end_point_vertical_error <= beta1:
            is_pass = True
else:
    is_pass = False
elif end_point_type == 1:  # 垂直
if end_point_horizontal_error <= alpha2 and
    end_point_vertical_error <= alpha1:
    is_pass = True
    else:
    is_pass = False
    else:
    is_pass = False
    else:  # 下一个点是终点
    if end_point_horizontal_error <= theta and end_point_vertical_error <= theta:
        is_pass = True
    else:
        is_pass = False
    else:
    is_pass = False
    after_end_point_horizontal_error = end_point_horizontal_error
    after_end_point_vertical_error = end_point_vertical_error
    if is_pass:
        if end_point_type == 0:
            after_end_point_horizontal_error = 0
    after_end_point_vertical_error = end_point_vertical_error
    elif end_point_type == 1:
    after_end_point_horizontal_error = end_point_horizontal_error
    after_end_point_vertical_error = 0
    return is_pass, end_point_horizontal_error, end_point_vertical_error,
    after_end_point_horizontal_error, after_end_point_vertical_error

    matlab_path_list = [0, 503, 294, 91, 282, 33, 315, 403, 594, 501, 612]
    path_list = [[i, 0, 0, 0, 0] for i in matlab_path_list]
    distance = 0
    for i in range(len(path_list) - 1):
        start_index = path_list[i][0]
        end_index = path_list[i + 1][0]

    is_pass, end_point_horizontal_error, end_point_vertical_error,
    after_end_point_horizontal_error, after_end_point_vertical_error = judge(start_index, end_index,
                                                                             path_list[i][3], path_list[i][4])
    if is_pass:
        path_list[i + 1][1:] = end_point_horizontal_error, end_point_vertical_error,
    after_end_point_horizontal_error, after_end_point_vertical_error
    else:
    print('error', start_index, end_index)
    distance = distance + get_distance(start_index, end_index)
    for i in path_list:
        print(i, point_list[i[0]][3])
